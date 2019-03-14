# -*- coding:utf-8 -*-
import openpyxl
import config.config as cf


def write_to_excel(path, sheetname, row, col, string):
    """ 将测试结果写入到Excel """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    # sheet.cell_value(rowx=row, colx=col).value = string
    # sheet.cell(row=row, col=col).value = string
    sheet.cell(row=row, column=col).value = string
    workbook.save(path)


if __name__ == '__main__':
    res = write_to_excel(cf.CASE_PATH, cf.CASE_SHEET_NAME, 2, cf.CASE_TEST_RESULT, 'pass')
